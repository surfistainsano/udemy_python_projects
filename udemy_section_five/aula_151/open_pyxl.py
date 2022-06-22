import openpyxl
from random import uniform

pedidos = openpyxl.load_workbook('pedidos.xlsx')  # A variável herda os métodos e propriedades do WorkBook
nome_planilhas = pedidos.sheetnames  # É exibido o nome das planilhas existentes no arquivo
planilha1 = pedidos['Página1']  # É possível selecionar a planilha como se fosse um dict

"""
É importante notar que o pycharm não dá sugestões para a variável ->planilha1<- ao fazer um acesso com o dot.
Por isso é importante ler a documentação, para ter conhecimento do que esse módulo é capaz de fazer.
Essa variável retorna um object Worksheet.
"""

# Para acessar a célula B4 da planilha:
# print(planilha1['b4'])  # // object Cell

cell_b4_value = planilha1['b4'].value  # retorna o valor da célula b4

column_b = planilha1['b']  # Retorna uma tupla; cada item é um object Cell representando uma linha.

"""
# Para consultar um range (intervalo) de células:
for linha in planilha1['a1:c2']:  # planilha1['a1:c2'] Retorna uma tupla contendo as células daquela linha
    # em outra tupla | // ((<Cell 'Página1'.A1>, <Cell 'Página1'.B1>, <Cell 'Página1'.C1>),
    # (<Cell 'Página1'.A2>, <Cell 'Página1'.B2>, <Cell 'Página1'.C2>))
    for coluna in linha:
        print(coluna.value)
"""

"""
# É possível percorrer a planilha inteira:
for linha in planilha1:
    # len(linha) Para saber a quantidade de colunas na linha
    if linha[0].value is not None:
        print(linha[0].value, end=" ")
    if linha[1].value is not None:
        print(linha[1].value, end=" ")
    if linha[2].value is not None:
        print(linha[2].value)
"""

planilha1['b3'].value = 3200  # Não altera a planilha original

"""
# Para salvar as alterações em um novo arquivo:
# pedidos.save('planilha_modificada.xlsx')

for linha in range(5, 16):  # o for percorrerá até a linha 15
    num_pedido = linha - 1  # 4 até 14
    planilha1.cell(linha, 1).value = num_pedido
    # object.cell(num_linha, num_coluna) A coluna é representada por número e não por letra
    planilha1.cell(linha, 2).value = 1200 + linha

    preco = round(uniform(10, 100), 2)
    planilha1.cell(linha, 3).value = preco

pedidos.save('planilha_modificada.xlsx')
"""

# Para criar uma nova planilha

planilha = openpyxl.Workbook()
planilha.create_sheet('Planilha1', 0)
planilha.create_sheet('Planilha2', 1)

planilha_1 = planilha['Planilha1']
planilha_2 = planilha['Planilha2']

for linha in range(1, 11):
    num_pedido = linha - 1
    planilha_1.cell(linha, 1).value = num_pedido
    planilha_1.cell(linha, 2).value = 1200 + linha

    preco = round(uniform(10, 100), 2)
    planilha_1.cell(linha, 3).value = preco

for linha in range(1, 11):
    planilha_2.cell(linha, 1).value = f'Leonardo {linha} {round(uniform(10, 100), 2)}'
    planilha_2.cell(linha, 2).value = f'Camila {linha} {round(uniform(10, 100), 2)}'
    planilha_2.cell(linha, 3).value = f'Thaís {linha} {round(uniform(10, 100), 2)}'

planilha.save('nova_planilha.xlsx')
